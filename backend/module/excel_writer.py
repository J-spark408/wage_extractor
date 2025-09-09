import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from .utils import ensure_dir, join

def write_excel(output_dir: str, employee_name: str, date_range: str, rows: list) -> str:
    ensure_dir(output_dir)
    file_name_date = date_range.replace("/", ".")
    path = join(output_dir, f"Wage Information Request_{file_name_date}.xlsx")

    # DataFrame with Week column first
    df = pd.DataFrame(rows)
    cols = df.columns.tolist()
    cols.insert(0, cols.pop(cols.index("Week")))
    df = df[cols]
    df.to_excel(path, index=False)

    # Center & widths
    wb = load_workbook(path)
    ws = wb.active
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in r:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for col, w in {"A": 10, "B": 15, "C": 15, "D": 12}.items():
        ws.column_dimensions[col].width = w
    wb.save(path)
    return path
